"""
Report Export Service

Provides comprehensive export functionality for reports in multiple formats:
- PDF: Professional reports with charts and styling
- Excel: Rich formatting with multiple sheets and charts
- CSV: Simple tabular data export
- JSON: Structured data export for API consumption
- HTML: Web-friendly report format
"""

import logging
import json
import io
import csv
from datetime import datetime, timezone, timedelta
from typing import Dict, List, Optional, Any, Union, Tuple
from pathlib import Path
import tempfile
import shutil

# Excel export
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference, PieChart
from openpyxl.utils.dataframe import dataframe_to_rows

# PDF export
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY

# Data processing
import pandas as pd
import numpy as np

from ..utils.export_utils import ExportManager
from .weekly_reports import WeeklyReportGenerator
from .report_analytics import ReportAnalytics
from .report_archiving import ReportArchiver, ReportType

logger = logging.getLogger(__name__)


class ReportExportService:
    """Comprehensive report export service for multiple formats and report types."""
    
    def __init__(self):
        self.export_manager = ExportManager()
        self.weekly_generator = WeeklyReportGenerator()
        self.analytics_service = ReportAnalytics()
        self.archiver = ReportArchiver()
        self.supported_formats = ['pdf', 'excel', 'csv', 'json', 'html']
        self.max_export_size = 50000  # Maximum records per export
        
    def export_weekly_report(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        format: str = 'pdf',
        include_charts: bool = True,
        include_analytics: bool = True,
        filters: Optional[Dict[str, Any]] = None,
        custom_title: Optional[str] = None
    ) -> Union[str, bytes]:
        """
        Export a weekly report in the specified format.
        
        Args:
            start_date: Start date for report period
            end_date: End date for report period
            format: Export format ('pdf', 'excel', 'csv', 'json', 'html')
            include_charts: Include charts and visualizations
            include_analytics: Include analytics and trend analysis
            filters: Additional filters to apply
            custom_title: Custom title for the report
            
        Returns:
            Report content in the specified format
        """
        if format not in self.supported_formats:
            raise ValueError(f"Unsupported format: {format}")
            
        # Generate the weekly report data
        report_data = self.weekly_generator.generate_weekly_report(
            start_date=start_date,
            end_date=end_date,
            states=filters.get('states') if filters else None,
            form_types=filters.get('form_types') if filters else None,
            severity_levels=filters.get('severity_levels') if filters else None,
            include_ai_analysis=True,
            include_impact_assessment=True,
            include_notification_summary=True,
            include_monitoring_statistics=True
        )
        
        # Add analytics if requested
        if include_analytics:
            analytics_data = self.analytics_service.generate_comprehensive_analytics(
                start_date=start_date,
                end_date=end_date,
                agencies=filters.get('agencies') if filters else None,
                form_types=filters.get('form_types') if filters else None
            )
            report_data['analytics'] = analytics_data
        
        # Set custom title if provided
        if custom_title:
            report_data['custom_title'] = custom_title
        
        # Export based on format
        if format == 'pdf':
            return self._export_weekly_pdf(report_data, include_charts)
        elif format == 'excel':
            return self._export_weekly_excel(report_data, include_charts)
        elif format == 'csv':
            return self._export_weekly_csv(report_data)
        elif format == 'json':
            return self._export_weekly_json(report_data)
        elif format == 'html':
            return self._export_weekly_html(report_data, include_charts)
    
    def export_analytics_report(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        format: str = 'pdf',
        include_predictions: bool = True,
        include_anomalies: bool = True,
        include_correlations: bool = True,
        agencies: Optional[List[int]] = None,
        form_types: Optional[List[str]] = None
    ) -> Union[str, bytes]:
        """
        Export an analytics report in the specified format.
        
        Args:
            start_date: Start date for analysis period
            end_date: End date for analysis period
            format: Export format
            include_predictions: Include predictive analytics
            include_anomalies: Include anomaly detection
            include_correlations: Include correlation analysis
            agencies: Filter by specific agencies
            form_types: Filter by specific form types
            
        Returns:
            Analytics report content in the specified format
        """
        if format not in self.supported_formats:
            raise ValueError(f"Unsupported format: {format}")
            
        # Generate analytics data
        analytics_data = self.analytics_service.generate_comprehensive_analytics(
            start_date=start_date,
            end_date=end_date,
            agencies=agencies,
            form_types=form_types,
            include_predictions=include_predictions,
            include_anomalies=include_anomalies,
            include_correlations=include_correlations
        )
        
        # Export based on format
        if format == 'pdf':
            return self._export_analytics_pdf(analytics_data)
        elif format == 'excel':
            return self._export_analytics_excel(analytics_data)
        elif format == 'csv':
            return self._export_analytics_csv(analytics_data)
        elif format == 'json':
            return self._export_analytics_json(analytics_data)
        elif format == 'html':
            return self._export_analytics_html(analytics_data)
    
    def export_archive_report(
        self,
        report_id: str,
        format: str = 'pdf',
        include_metadata: bool = True
    ) -> Union[str, bytes]:
        """
        Export an archived report in the specified format.
        
        Args:
            report_id: ID of the archived report
            format: Export format
            include_metadata: Include archive metadata
            
        Returns:
            Archived report content in the specified format
        """
        if format not in self.supported_formats:
            raise ValueError(f"Unsupported format: {format}")
            
        # Retrieve the archived report
        archived_data = self.archiver.retrieve_report(report_id)
        if not archived_data:
            raise ValueError(f"Archived report {report_id} not found")
            
        # Get metadata if requested
        metadata = None
        if include_metadata:
            metadata = self.archiver._get_metadata(report_id)
        
        # Export based on format
        if format == 'pdf':
            return self._export_archive_pdf(archived_data, metadata)
        elif format == 'excel':
            return self._export_archive_excel(archived_data, metadata)
        elif format == 'csv':
            return self._export_archive_csv(archived_data)
        elif format == 'json':
            return self._export_archive_json(archived_data, metadata)
        elif format == 'html':
            return self._export_archive_html(archived_data, metadata)
    
    def export_custom_report(
        self,
        report_data: Dict[str, Any],
        format: str = 'pdf',
        report_type: str = 'custom',
        include_charts: bool = True
    ) -> Union[str, bytes]:
        """
        Export a custom report in the specified format.
        
        Args:
            report_data: Custom report data
            format: Export format
            report_type: Type of report for metadata
            include_charts: Include charts and visualizations
            
        Returns:
            Custom report content in the specified format
        """
        if format not in self.supported_formats:
            raise ValueError(f"Unsupported format: {format}")
            
        # Export based on format
        if format == 'pdf':
            return self._export_custom_pdf(report_data, report_type, include_charts)
        elif format == 'excel':
            return self._export_custom_excel(report_data, report_type, include_charts)
        elif format == 'csv':
            return self._export_custom_csv(report_data)
        elif format == 'json':
            return self._export_custom_json(report_data)
        elif format == 'html':
            return self._export_custom_html(report_data, report_type, include_charts)
    
    def _export_weekly_pdf(self, report_data: Dict[str, Any], include_charts: bool) -> bytes:
        """Export weekly report to PDF format."""
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title = report_data.get('custom_title', 'Weekly Compliance Report')
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 20))
        
        # Executive Summary
        if 'executive_summary' in report_data:
            story.extend(self._create_pdf_executive_summary(report_data['executive_summary'], styles))
            story.append(Spacer(1, 20))
        
        # Form Changes
        if 'form_changes' in report_data:
            story.extend(self._create_pdf_form_changes(report_data['form_changes'], styles))
            story.append(Spacer(1, 20))
        
        # Analytics
        if 'analytics' in report_data:
            story.extend(self._create_pdf_analytics_summary(report_data['analytics'], styles))
            story.append(Spacer(1, 20))
        
        # Monitoring Statistics
        if 'monitoring_statistics' in report_data:
            story.extend(self._create_pdf_monitoring_stats(report_data['monitoring_statistics'], styles))
            story.append(Spacer(1, 20))
        
        # Recommendations
        if 'recommendations' in report_data:
            story.extend(self._create_pdf_recommendations(report_data['recommendations'], styles))
        
        doc.build(story)
        return output.getvalue()
    
    def _export_weekly_excel(self, report_data: Dict[str, Any], include_charts: bool) -> bytes:
        """Export weekly report to Excel format."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Executive Summary sheet
        self._create_excel_executive_summary(wb, report_data)
        
        # Form Changes sheet
        if 'form_changes' in report_data:
            self._create_excel_form_changes(wb, report_data['form_changes'])
        
        # Analytics sheet
        if 'analytics' in report_data:
            self._create_excel_analytics(wb, report_data['analytics'])
        
        # Monitoring Statistics sheet
        if 'monitoring_statistics' in report_data:
            self._create_excel_monitoring_stats(wb, report_data['monitoring_statistics'])
        
        # Charts sheet
        if include_charts and 'form_changes' in report_data:
            self._create_excel_charts(wb, report_data['form_changes'])
        
        # Metadata sheet
        self._create_excel_metadata(wb, report_data)
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _export_weekly_csv(self, report_data: Dict[str, Any]) -> str:
        """Export weekly report to CSV format."""
        output = io.StringIO()
        
        # Export form changes as CSV
        if 'form_changes' in report_data and report_data['form_changes']:
            changes = report_data['form_changes']
            if changes:
                writer = csv.DictWriter(output, fieldnames=changes[0].keys())
                writer.writeheader()
                for change in changes:
                    writer.writerow(self._format_row_for_csv(change))
        
        return output.getvalue()
    
    def _export_weekly_json(self, report_data: Dict[str, Any]) -> str:
        """Export weekly report to JSON format."""
        return json.dumps(report_data, default=str, indent=2)
    
    def _export_weekly_html(self, report_data: Dict[str, Any], include_charts: bool) -> str:
        """Export weekly report to HTML format."""
        title = report_data.get('custom_title', 'Weekly Compliance Report')
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{title}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ text-align: center; margin-bottom: 30px; }}
                .section {{ margin-bottom: 30px; }}
                .section h2 {{ color: #333; border-bottom: 2px solid #007bff; }}
                table {{ width: 100%; border-collapse: collapse; margin: 10px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                .summary {{ background-color: #f8f9fa; padding: 15px; border-radius: 5px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{title}</h1>
                <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
        """
        
        # Executive Summary
        if 'executive_summary' in report_data:
            html += self._create_html_executive_summary(report_data['executive_summary'])
        
        # Form Changes
        if 'form_changes' in report_data:
            html += self._create_html_form_changes(report_data['form_changes'])
        
        # Analytics
        if 'analytics' in report_data:
            html += self._create_html_analytics_summary(report_data['analytics'])
        
        html += """
        </body>
        </html>
        """
        
        return html
    
    def _create_pdf_executive_summary(self, summary: Dict[str, Any], styles) -> List:
        """Create PDF executive summary section."""
        story = []
        
        story.append(Paragraph("Executive Summary", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Summary statistics
        if 'total_changes' in summary:
            story.append(Paragraph(f"Total Changes: {summary['total_changes']}", styles['Normal']))
        if 'critical_changes' in summary:
            story.append(Paragraph(f"Critical Changes: {summary['critical_changes']}", styles['Normal']))
        if 'high_priority_changes' in summary:
            story.append(Paragraph(f"High Priority Changes: {summary['high_priority_changes']}", styles['Normal']))
        
        story.append(Spacer(1, 12))
        
        # Key insights
        if 'key_insights' in summary:
            story.append(Paragraph("Key Insights:", styles['Heading3']))
            for insight in summary['key_insights']:
                story.append(Paragraph(f"• {insight}", styles['Normal']))
        
        return story
    
    def _create_pdf_form_changes(self, changes: List[Dict[str, Any]], styles) -> List:
        """Create PDF form changes section."""
        story = []
        
        story.append(Paragraph("Form Changes", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        if not changes:
            story.append(Paragraph("No changes detected in this period.", styles['Normal']))
            return story
        
        # Create table
        headers = ['Agency', 'Form', 'Change Type', 'Severity', 'Date']
        table_data = [headers]
        
        for change in changes[:50]:  # Limit to first 50 for PDF
            row = [
                change.get('agency_name', ''),
                change.get('form_name', ''),
                change.get('change_type', ''),
                change.get('severity', ''),
                change.get('detected_at', '')
            ]
            table_data.append(row)
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(table)
        
        if len(changes) > 50:
            story.append(Paragraph(f"Note: Showing first 50 of {len(changes)} changes", styles['Normal']))
        
        return story
    
    def _create_excel_executive_summary(self, wb: openpyxl.Workbook, report_data: Dict[str, Any]):
        """Create Excel executive summary sheet."""
        ws = wb.create_sheet("Executive Summary")
        
        # Title
        title = report_data.get('custom_title', 'Weekly Compliance Report')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        
        # Summary data
        if 'executive_summary' in report_data:
            summary = report_data['executive_summary']
            row = 3
            
            if 'total_changes' in summary:
                ws[f'A{row}'] = "Total Changes:"
                ws[f'B{row}'] = summary['total_changes']
                row += 1
            
            if 'critical_changes' in summary:
                ws[f'A{row}'] = "Critical Changes:"
                ws[f'B{row}'] = summary['critical_changes']
                row += 1
            
            if 'high_priority_changes' in summary:
                ws[f'A{row}'] = "High Priority Changes:"
                ws[f'B{row}'] = summary['high_priority_changes']
                row += 1
    
    def _create_excel_form_changes(self, wb: openpyxl.Workbook, changes: List[Dict[str, Any]]):
        """Create Excel form changes sheet."""
        ws = wb.create_sheet("Form Changes")
        
        if not changes:
            ws['A1'] = "No changes detected in this period."
            return
        
        # Headers
        headers = ['Agency', 'Form', 'Change Type', 'Severity', 'Date', 'Description']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data
        for row_idx, change in enumerate(changes, 2):
            ws.cell(row=row_idx, column=1, value=change.get('agency_name', ''))
            ws.cell(row=row_idx, column=2, value=change.get('form_name', ''))
            ws.cell(row=row_idx, column=3, value=change.get('change_type', ''))
            ws.cell(row=row_idx, column=4, value=change.get('severity', ''))
            ws.cell(row=row_idx, column=5, value=change.get('detected_at', ''))
            ws.cell(row=row_idx, column=6, value=change.get('description', ''))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _format_row_for_csv(self, row: Dict[str, Any]) -> Dict[str, Any]:
        """Format row data for CSV export."""
        formatted_row = {}
        for key, value in row.items():
            if isinstance(value, datetime):
                formatted_row[key] = value.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(value, dict):
                formatted_row[key] = json.dumps(value)
            else:
                formatted_row[key] = str(value) if value is not None else ""
        return formatted_row
    
    def _export_analytics_pdf(self, analytics_data: Dict[str, Any]) -> bytes:
        """Export analytics report to PDF format."""
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph("Analytics Report", title_style))
        story.append(Spacer(1, 20))
        
        # Summary
        if 'summary' in analytics_data:
            story.extend(self._create_pdf_analytics_summary(analytics_data['summary'], styles))
            story.append(Spacer(1, 20))
        
        # Trend Analysis
        if 'trend_analysis' in analytics_data:
            story.extend(self._create_pdf_trend_analysis(analytics_data['trend_analysis'], styles))
            story.append(Spacer(1, 20))
        
        # Insights
        if 'insights' in analytics_data:
            story.extend(self._create_pdf_insights(analytics_data['insights'], styles))
        
        doc.build(story)
        return output.getvalue()
    
    def _export_analytics_excel(self, analytics_data: Dict[str, Any]) -> bytes:
        """Export analytics report to Excel format."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Summary sheet
        self._create_excel_analytics_summary(wb, analytics_data)
        
        # Trend Analysis sheet
        if 'trend_analysis' in analytics_data:
            self._create_excel_trend_analysis(wb, analytics_data['trend_analysis'])
        
        # Insights sheet
        if 'insights' in analytics_data:
            self._create_excel_insights(wb, analytics_data['insights'])
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _export_analytics_csv(self, analytics_data: Dict[str, Any]) -> str:
        """Export analytics report to CSV format."""
        output = io.StringIO()
        
        # Export trend data as CSV
        if 'trend_analysis' in analytics_data and 'daily_changes' in analytics_data['trend_analysis']:
            changes = analytics_data['trend_analysis']['daily_changes']
            if changes:
                writer = csv.DictWriter(output, fieldnames=changes[0].keys())
                writer.writeheader()
                for change in changes:
                    writer.writerow(self._format_row_for_csv(change))
        
        return output.getvalue()
    
    def _export_analytics_json(self, analytics_data: Dict[str, Any]) -> str:
        """Export analytics report to JSON format."""
        return json.dumps(analytics_data, default=str, indent=2)
    
    def _export_analytics_html(self, analytics_data: Dict[str, Any]) -> str:
        """Export analytics report to HTML format."""
        html = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Analytics Report</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                .header { text-align: center; margin-bottom: 30px; }
                .section { margin-bottom: 30px; }
                .section h2 { color: #333; border-bottom: 2px solid #007bff; }
                table { width: 100%; border-collapse: collapse; margin: 10px 0; }
                th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                th { background-color: #f2f2f2; }
                .summary { background-color: #f8f9fa; padding: 15px; border-radius: 5px; }
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Analytics Report</h1>
                <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
        """
        
        # Summary
        if 'summary' in analytics_data:
            html += self._create_html_analytics_summary(analytics_data['summary'])
        
        # Trend Analysis
        if 'trend_analysis' in analytics_data:
            html += self._create_html_trend_analysis(analytics_data['trend_analysis'])
        
        html += """
        </body>
        </html>
        """
        
        return html
    
    def _export_archive_pdf(self, archived_data: Dict[str, Any], metadata: Optional[Dict[str, Any]]) -> bytes:
        """Export archived report to PDF format."""
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title = "Archived Report"
        if metadata and 'title' in metadata:
            title = metadata['title']
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 20))
        
        # Metadata
        if metadata:
            story.extend(self._create_pdf_archive_metadata(metadata, styles))
            story.append(Spacer(1, 20))
        
        # Report content
        story.extend(self._create_pdf_archive_content(archived_data, styles))
        
        doc.build(story)
        return output.getvalue()
    
    def _export_archive_excel(self, archived_data: Dict[str, Any], metadata: Optional[Dict[str, Any]]) -> bytes:
        """Export archived report to Excel format."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Metadata sheet
        if metadata:
            self._create_excel_archive_metadata(wb, metadata)
        
        # Content sheet
        self._create_excel_archive_content(wb, archived_data)
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _export_archive_csv(self, archived_data: Dict[str, Any]) -> str:
        """Export archived report to CSV format."""
        output = io.StringIO()
        
        # Export form changes as CSV
        if 'form_changes' in archived_data and archived_data['form_changes']:
            changes = archived_data['form_changes']
            if changes:
                writer = csv.DictWriter(output, fieldnames=changes[0].keys())
                writer.writeheader()
                for change in changes:
                    writer.writerow(self._format_row_for_csv(change))
        
        return output.getvalue()
    
    def _export_archive_json(self, archived_data: Dict[str, Any], metadata: Optional[Dict[str, Any]]) -> str:
        """Export archived report to JSON format."""
        export_data = {
            'archived_data': archived_data,
            'metadata': metadata
        }
        return json.dumps(export_data, default=str, indent=2)
    
    def _export_archive_html(self, archived_data: Dict[str, Any], metadata: Optional[Dict[str, Any]]) -> str:
        """Export archived report to HTML format."""
        title = "Archived Report"
        if metadata and 'title' in metadata:
            title = metadata['title']
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{title}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ text-align: center; margin-bottom: 30px; }}
                .section {{ margin-bottom: 30px; }}
                .section h2 {{ color: #333; border-bottom: 2px solid #007bff; }}
                table {{ width: 100%; border-collapse: collapse; margin: 10px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                .summary {{ background-color: #f8f9fa; padding: 15px; border-radius: 5px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{title}</h1>
                <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
        """
        
        # Metadata
        if metadata:
            html += self._create_html_archive_metadata(metadata)
        
        # Content
        html += self._create_html_archive_content(archived_data)
        
        html += """
        </body>
        </html>
        """
        
        return html
    
    def _export_custom_pdf(self, report_data: Dict[str, Any], report_type: str, include_charts: bool) -> bytes:
        """Export custom report to PDF format."""
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title = report_data.get('title', f'{report_type.title()} Report')
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 20))
        
        # Content
        story.extend(self._create_pdf_custom_content(report_data, styles))
        
        doc.build(story)
        return output.getvalue()
    
    def _export_custom_excel(self, report_data: Dict[str, Any], report_type: str, include_charts: bool) -> bytes:
        """Export custom report to Excel format."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Content sheet
        self._create_excel_custom_content(wb, report_data, report_type)
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _export_custom_csv(self, report_data: Dict[str, Any]) -> str:
        """Export custom report to CSV format."""
        output = io.StringIO()
        
        # Export data as CSV
        if 'data' in report_data and report_data['data']:
            data = report_data['data']
            if data:
                writer = csv.DictWriter(output, fieldnames=data[0].keys())
                writer.writeheader()
                for row in data:
                    writer.writerow(self._format_row_for_csv(row))
        
        return output.getvalue()
    
    def _export_custom_json(self, report_data: Dict[str, Any]) -> str:
        """Export custom report to JSON format."""
        return json.dumps(report_data, default=str, indent=2)
    
    def _export_custom_html(self, report_data: Dict[str, Any], report_type: str, include_charts: bool) -> str:
        """Export custom report to HTML format."""
        title = report_data.get('title', f'{report_type.title()} Report')
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{title}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ text-align: center; margin-bottom: 30px; }}
                .section {{ margin-bottom: 30px; }}
                .section h2 {{ color: #333; border-bottom: 2px solid #007bff; }}
                table {{ width: 100%; border-collapse: collapse; margin: 10px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                .summary {{ background-color: #f8f9fa; padding: 15px; border-radius: 5px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{title}</h1>
                <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
        """
        
        # Content
        html += self._create_html_custom_content(report_data)
        
        html += """
        </body>
        </html>
        """
        
        return html
    
    # Helper methods for PDF creation
    def _create_pdf_analytics_summary(self, summary: Dict[str, Any], styles) -> List:
        """Create PDF analytics summary section."""
        story = []
        
        story.append(Paragraph("Analytics Summary", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Summary statistics
        if 'data_points_analyzed' in summary:
            story.append(Paragraph(f"Data Points Analyzed: {summary['data_points_analyzed']}", styles['Normal']))
        if 'agencies_analyzed' in summary:
            story.append(Paragraph(f"Agencies Analyzed: {summary['agencies_analyzed']}", styles['Normal']))
        
        return story
    
    def _create_pdf_trend_analysis(self, trend_data: Dict[str, Any], styles) -> List:
        """Create PDF trend analysis section."""
        story = []
        
        story.append(Paragraph("Trend Analysis", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Trend information
        if 'trend_direction' in trend_data:
            story.append(Paragraph(f"Trend Direction: {trend_data['trend_direction']}", styles['Normal']))
        if 'trend_strength' in trend_data:
            story.append(Paragraph(f"Trend Strength: {trend_data['trend_strength']}", styles['Normal']))
        
        return story
    
    def _create_pdf_insights(self, insights: List[str], styles) -> List:
        """Create PDF insights section."""
        story = []
        
        story.append(Paragraph("Key Insights", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        for insight in insights:
            story.append(Paragraph(f"• {insight}", styles['Normal']))
        
        return story
    
    def _create_pdf_monitoring_stats(self, stats: Dict[str, Any], styles) -> List:
        """Create PDF monitoring statistics section."""
        story = []
        
        story.append(Paragraph("Monitoring Statistics", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        if 'total_runs' in stats:
            story.append(Paragraph(f"Total Monitoring Runs: {stats['total_runs']}", styles['Normal']))
        if 'success_rate' in stats:
            story.append(Paragraph(f"Success Rate: {stats['success_rate']}%", styles['Normal']))
        
        return story
    
    def _create_pdf_recommendations(self, recommendations: List[str], styles) -> List:
        """Create PDF recommendations section."""
        story = []
        
        story.append(Paragraph("Recommendations", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        for recommendation in recommendations:
            story.append(Paragraph(f"• {recommendation}", styles['Normal']))
        
        return story
    
    def _create_pdf_archive_metadata(self, metadata: Dict[str, Any], styles) -> List:
        """Create PDF archive metadata section."""
        story = []
        
        story.append(Paragraph("Archive Metadata", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        if 'generated_at' in metadata:
            story.append(Paragraph(f"Generated: {metadata['generated_at']}", styles['Normal']))
        if 'report_type' in metadata:
            story.append(Paragraph(f"Report Type: {metadata['report_type']}", styles['Normal']))
        
        return story
    
    def _create_pdf_archive_content(self, archived_data: Dict[str, Any], styles) -> List:
        """Create PDF archive content section."""
        story = []
        
        story.append(Paragraph("Report Content", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Add content based on what's available
        if 'form_changes' in archived_data:
            story.extend(self._create_pdf_form_changes(archived_data['form_changes'], styles))
        
        return story
    
    def _create_pdf_custom_content(self, report_data: Dict[str, Any], styles) -> List:
        """Create PDF custom content section."""
        story = []
        
        # Add content based on what's available
        if 'content' in report_data:
            story.append(Paragraph(report_data['content'], styles['Normal']))
        
        return story
    
    # Helper methods for Excel creation
    def _create_excel_analytics(self, wb: openpyxl.Workbook, analytics_data: Dict[str, Any]):
        """Create Excel analytics sheet."""
        ws = wb.create_sheet("Analytics")
        
        # Add analytics data
        if 'summary' in analytics_data:
            summary = analytics_data['summary']
            ws['A1'] = "Analytics Summary"
            ws['A1'].font = Font(bold=True, size=16)
            
            row = 3
            if 'data_points_analyzed' in summary:
                ws[f'A{row}'] = "Data Points Analyzed:"
                ws[f'B{row}'] = summary['data_points_analyzed']
                row += 1
    
    def _create_excel_monitoring_stats(self, wb: openpyxl.Workbook, stats: Dict[str, Any]):
        """Create Excel monitoring statistics sheet."""
        ws = wb.create_sheet("Monitoring Statistics")
        
        ws['A1'] = "Monitoring Statistics"
        ws['A1'].font = Font(bold=True, size=16)
        
        row = 3
        if 'total_runs' in stats:
            ws[f'A{row}'] = "Total Runs:"
            ws[f'B{row}'] = stats['total_runs']
            row += 1
    
    def _create_excel_charts(self, wb: openpyxl.Workbook, changes: List[Dict[str, Any]]):
        """Create Excel charts sheet."""
        ws = wb.create_sheet("Charts")
        
        # Create severity pie chart
        severity_counts = {}
        for change in changes:
            severity = change.get('severity', 'Unknown')
            severity_counts[severity] = severity_counts.get(severity, 0) + 1
        
        if severity_counts:
            # Add chart data
            ws['A1'] = "Severity"
            ws['B1'] = "Count"
            row = 2
            for severity, count in severity_counts.items():
                ws[f'A{row}'] = severity
                ws[f'B{row}'] = count
                row += 1
            
            # Create pie chart
            chart = PieChart()
            chart.title = "Changes by Severity"
            data = Reference(ws, min_col=2, min_row=1, max_row=len(severity_counts) + 1)
            labels = Reference(ws, min_col=1, min_row=2, max_row=len(severity_counts) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            ws.add_chart(chart, "D2")
    
    def _create_excel_metadata(self, wb: openpyxl.Workbook, report_data: Dict[str, Any]):
        """Create Excel metadata sheet."""
        ws = wb.create_sheet("Metadata")
        
        ws['A1'] = "Report Metadata"
        ws['A1'].font = Font(bold=True, size=16)
        
        ws['A3'] = "Generated:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws['A4'] = "Format:"
        ws['B4'] = "Excel"
    
    def _create_excel_analytics_summary(self, wb: openpyxl.Workbook, analytics_data: Dict[str, Any]):
        """Create Excel analytics summary sheet."""
        ws = wb.create_sheet("Analytics Summary")
        
        ws['A1'] = "Analytics Summary"
        ws['A1'].font = Font(bold=True, size=16)
        
        if 'summary' in analytics_data:
            summary = analytics_data['summary']
            row = 3
            if 'data_points_analyzed' in summary:
                ws[f'A{row}'] = "Data Points Analyzed:"
                ws[f'B{row}'] = summary['data_points_analyzed']
                row += 1
    
    def _create_excel_trend_analysis(self, wb: openpyxl.Workbook, trend_data: Dict[str, Any]):
        """Create Excel trend analysis sheet."""
        ws = wb.create_sheet("Trend Analysis")
        
        ws['A1'] = "Trend Analysis"
        ws['A1'].font = Font(bold=True, size=16)
        
        row = 3
        if 'trend_direction' in trend_data:
            ws[f'A{row}'] = "Trend Direction:"
            ws[f'B{row}'] = trend_data['trend_direction']
            row += 1
    
    def _create_excel_insights(self, wb: openpyxl.Workbook, insights: List[str]):
        """Create Excel insights sheet."""
        ws = wb.create_sheet("Insights")
        
        ws['A1'] = "Key Insights"
        ws['A1'].font = Font(bold=True, size=16)
        
        for i, insight in enumerate(insights, 2):
            ws[f'A{i}'] = f"• {insight}"
    
    def _create_excel_archive_metadata(self, wb: openpyxl.Workbook, metadata: Dict[str, Any]):
        """Create Excel archive metadata sheet."""
        ws = wb.create_sheet("Archive Metadata")
        
        ws['A1'] = "Archive Metadata"
        ws['A1'].font = Font(bold=True, size=16)
        
        row = 3
        if 'generated_at' in metadata:
            ws[f'A{row}'] = "Generated:"
            ws[f'B{row}'] = metadata['generated_at']
            row += 1
    
    def _create_excel_archive_content(self, wb: openpyxl.Workbook, archived_data: Dict[str, Any]):
        """Create Excel archive content sheet."""
        ws = wb.create_sheet("Report Content")
        
        ws['A1'] = "Report Content"
        ws['A1'].font = Font(bold=True, size=16)
        
        if 'form_changes' in archived_data:
            changes = archived_data['form_changes']
            if changes:
                headers = ['Agency', 'Form', 'Change Type', 'Severity', 'Date']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col, value=header)
                    cell.font = Font(bold=True)
                
                for row_idx, change in enumerate(changes, 4):
                    ws.cell(row=row_idx, column=1, value=change.get('agency_name', ''))
                    ws.cell(row=row_idx, column=2, value=change.get('form_name', ''))
                    ws.cell(row=row_idx, column=3, value=change.get('change_type', ''))
                    ws.cell(row=row_idx, column=4, value=change.get('severity', ''))
                    ws.cell(row=row_idx, column=5, value=change.get('detected_at', ''))
    
    def _create_excel_custom_content(self, wb: openpyxl.Workbook, report_data: Dict[str, Any], report_type: str):
        """Create Excel custom content sheet."""
        ws = wb.create_sheet("Content")
        
        title = report_data.get('title', f'{report_type.title()} Report')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        
        if 'content' in report_data:
            ws['A3'] = report_data['content']
    
    # Helper methods for HTML creation
    def _create_html_executive_summary(self, summary: Dict[str, Any]) -> str:
        """Create HTML executive summary section."""
        html = '<div class="section">'
        html += '<h2>Executive Summary</h2>'
        html += '<div class="summary">'
        
        if 'total_changes' in summary:
            html += f'<p><strong>Total Changes:</strong> {summary["total_changes"]}</p>'
        if 'critical_changes' in summary:
            html += f'<p><strong>Critical Changes:</strong> {summary["critical_changes"]}</p>'
        
        html += '</div></div>'
        return html
    
    def _create_html_form_changes(self, changes: List[Dict[str, Any]]) -> str:
        """Create HTML form changes section."""
        html = '<div class="section">'
        html += '<h2>Form Changes</h2>'
        
        if not changes:
            html += '<p>No changes detected in this period.</p>'
            html += '</div>'
            return html
        
        html += '<table>'
        html += '<tr><th>Agency</th><th>Form</th><th>Change Type</th><th>Severity</th><th>Date</th></tr>'
        
        for change in changes[:50]:  # Limit to first 50
            html += f'<tr>'
            html += f'<td>{change.get("agency_name", "")}</td>'
            html += f'<td>{change.get("form_name", "")}</td>'
            html += f'<td>{change.get("change_type", "")}</td>'
            html += f'<td>{change.get("severity", "")}</td>'
            html += f'<td>{change.get("detected_at", "")}</td>'
            html += f'</tr>'
        
        html += '</table>'
        
        if len(changes) > 50:
            html += f'<p><em>Note: Showing first 50 of {len(changes)} changes</em></p>'
        
        html += '</div>'
        return html
    
    def _create_html_analytics_summary(self, summary: Dict[str, Any]) -> str:
        """Create HTML analytics summary section."""
        html = '<div class="section">'
        html += '<h2>Analytics Summary</h2>'
        html += '<div class="summary">'
        
        if 'data_points_analyzed' in summary:
            html += f'<p><strong>Data Points Analyzed:</strong> {summary["data_points_analyzed"]}</p>'
        if 'agencies_analyzed' in summary:
            html += f'<p><strong>Agencies Analyzed:</strong> {summary["agencies_analyzed"]}</p>'
        
        html += '</div></div>'
        return html
    
    def _create_html_trend_analysis(self, trend_data: Dict[str, Any]) -> str:
        """Create HTML trend analysis section."""
        html = '<div class="section">'
        html += '<h2>Trend Analysis</h2>'
        html += '<div class="summary">'
        
        if 'trend_direction' in trend_data:
            html += f'<p><strong>Trend Direction:</strong> {trend_data["trend_direction"]}</p>'
        if 'trend_strength' in trend_data:
            html += f'<p><strong>Trend Strength:</strong> {trend_data["trend_strength"]}</p>'
        
        html += '</div></div>'
        return html
    
    def _create_html_archive_metadata(self, metadata: Dict[str, Any]) -> str:
        """Create HTML archive metadata section."""
        html = '<div class="section">'
        html += '<h2>Archive Metadata</h2>'
        html += '<div class="summary">'
        
        if 'generated_at' in metadata:
            html += f'<p><strong>Generated:</strong> {metadata["generated_at"]}</p>'
        if 'report_type' in metadata:
            html += f'<p><strong>Report Type:</strong> {metadata["report_type"]}</p>'
        
        html += '</div></div>'
        return html
    
    def _create_html_archive_content(self, archived_data: Dict[str, Any]) -> str:
        """Create HTML archive content section."""
        html = '<div class="section">'
        html += '<h2>Report Content</h2>'
        
        if 'form_changes' in archived_data:
            html += self._create_html_form_changes(archived_data['form_changes'])
        
        html += '</div>'
        return html
    
    def _create_html_custom_content(self, report_data: Dict[str, Any]) -> str:
        """Create HTML custom content section."""
        html = '<div class="section">'
        html += '<h2>Content</h2>'
        
        if 'content' in report_data:
            html += f'<p>{report_data["content"]}</p>'
        
        html += '</div>'
        return html


# Global export service instance
export_service = ReportExportService()


def export_weekly_report(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    format: str = 'pdf',
    **kwargs
) -> Union[str, bytes]:
    """Convenience function to export a weekly report."""
    return export_service.export_weekly_report(
        start_date=start_date,
        end_date=end_date,
        format=format,
        **kwargs
    )


def export_analytics_report(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    format: str = 'pdf',
    **kwargs
) -> Union[str, bytes]:
    """Convenience function to export an analytics report."""
    return export_service.export_analytics_report(
        start_date=start_date,
        end_date=end_date,
        format=format,
        **kwargs
    )


def export_archive_report(
    report_id: str,
    format: str = 'pdf',
    **kwargs
) -> Union[str, bytes]:
    """Convenience function to export an archived report."""
    return export_service.export_archive_report(
        report_id=report_id,
        format=format,
        **kwargs
    ) 