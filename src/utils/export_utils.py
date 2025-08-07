"""
Export utilities for dashboard data

Provides functionality to export filtered dashboard data in multiple formats:
- CSV: Simple tabular data export
- Excel: Rich formatting with multiple sheets and charts
- PDF: Professional reports with charts and styling
"""

import csv
import io
import json
from datetime import datetime, timezone, timedelta
from typing import List, Dict, Any, Optional, Union
from pathlib import Path
import logging

# Excel export
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

# PDF export
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Data processing
import pandas as pd
import numpy as np

logger = logging.getLogger(__name__)


class ExportManager:
    """Manages export operations for dashboard data."""
    
    def __init__(self):
        self.supported_formats = ['csv', 'excel', 'pdf']
        self.max_export_size = 10000  # Maximum records per export
        self.export_metadata = {
            'csv': {
                'mime_type': 'text/csv',
                'extension': '.csv',
                'description': 'Comma-separated values format for spreadsheet applications'
            },
            'excel': {
                'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'extension': '.xlsx', 
                'description': 'Microsoft Excel format with charts and formatting'
            },
            'pdf': {
                'mime_type': 'application/pdf',
                'extension': '.pdf',
                'description': 'Portable Document Format for professional reports'
            }
        }
        
    def export_data(
        self,
        data: List[Dict[str, Any]],
        format_type: str,
        export_config: Dict[str, Any],
        filename: Optional[str] = None
    ) -> Union[str, bytes]:
        """
        Export data in the specified format.
        
        Args:
            data: List of dictionaries containing the data to export
            format_type: Export format ('csv', 'excel', 'pdf')
            export_config: Configuration for the export (filters, columns, etc.)
            filename: Optional filename for the export
            
        Returns:
            File content as string (CSV) or bytes (Excel, PDF)
        """
        if format_type not in self.supported_formats:
            raise ValueError(f"Unsupported format: {format_type}")
        
        # Validate data and configuration
        self.validate_export_data(data, export_config)
            
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"compliance_export_{timestamp}.{format_type}"
            
        try:
            if format_type == 'csv':
                return self._export_csv(data, export_config, filename)
            elif format_type == 'excel':
                return self._export_excel(data, export_config, filename)
            elif format_type == 'pdf':
                return self._export_pdf(data, export_config, filename)
        except Exception as e:
            logger.error(f"Export failed for format {format_type}: {str(e)}")
            raise
    
    def get_export_metadata(self, format_type: str) -> Dict[str, Any]:
        """
        Get metadata for a specific export format.
        
        Args:
            format_type: Export format ('csv', 'excel', 'pdf')
            
        Returns:
            Dictionary containing format metadata
        """
        if format_type not in self.supported_formats:
            raise ValueError(f"Unsupported format: {format_type}")
        return self.export_metadata[format_type].copy()
    
    def get_supported_formats(self) -> List[Dict[str, Any]]:
        """
        Get list of all supported export formats with metadata.
        
        Returns:
            List of format dictionaries with metadata
        """
        formats = []
        for format_type in self.supported_formats:
            format_info = self.export_metadata[format_type].copy()
            format_info['format'] = format_type
            formats.append(format_info)
        return formats
    
    def validate_export_data(self, data: List[Dict[str, Any]], config: Dict[str, Any]) -> bool:
        """
        Validate export data and configuration before processing.
        
        Args:
            data: List of dictionaries containing the data to export
            config: Export configuration
            
        Returns:
            True if data is valid for export
            
        Raises:
            ValueError: If data or config is invalid
        """
        # Check data size
        if len(data) > self.max_export_size:
            raise ValueError(f"Export size exceeds maximum limit of {self.max_export_size} records")
        
        # Check columns configuration
        if config.get('columns') and data:
            available_columns = set(data[0].keys()) if data else set()
            requested_columns = set(config['columns'])
            missing_columns = requested_columns - available_columns
            
            if missing_columns:
                raise ValueError(f"Requested columns not found in data: {missing_columns}")
        
        return True
    
    def _export_csv(self, data: List[Dict[str, Any]], config: Dict[str, Any], filename: str) -> str:
        """Export data to CSV format."""
        if not data:
            return ""
            
        # Get column configuration
        columns = config.get('columns', list(data[0].keys()) if data else [])
        include_headers = config.get('include_headers', True)
        
        # Create CSV output
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns, extrasaction='ignore')
        
        if include_headers:
            writer.writeheader()
            
        for row in data:
            # Format datetime fields
            formatted_row = self._format_row_for_csv(row)
            writer.writerow(formatted_row)
            
        return output.getvalue()
    
    def _export_excel(self, data: List[Dict[str, Any]], config: Dict[str, Any], filename: str) -> bytes:
        """Export data to Excel format with formatting and charts."""
        if not data:
            return b""
            
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Export main data
        self._create_data_sheet(wb, data, config)
        
        # Export summary sheet
        self._create_summary_sheet(wb, data, config)
        
        # Export charts sheet
        self._create_charts_sheet(wb, data, config)
        
        # Export metadata
        self._create_metadata_sheet(wb, config)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _export_pdf(self, data: List[Dict[str, Any]], config: Dict[str, Any], filename: str) -> bytes:
        """Export data to PDF format with professional styling."""
        if not data:
            return b""
            
        # Create PDF document
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        
        # Build PDF content
        story = []
        styles = getSampleStyleSheet()
        
        # Add title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        title = Paragraph("Compliance Monitoring Report", title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Add export metadata
        story.extend(self._create_pdf_metadata(styles, config))
        story.append(Spacer(1, 20))
        
        # Add summary statistics
        story.extend(self._create_pdf_summary(data, styles))
        story.append(Spacer(1, 20))
        
        # Add data table
        story.extend(self._create_pdf_data_table(data, config, styles))
        
        # Build PDF
        doc.build(story)
        return output.getvalue()
    
    def _create_data_sheet(self, wb: openpyxl.Workbook, data: List[Dict[str, Any]], config: Dict[str, Any]):
        """Create the main data sheet in Excel."""
        ws = wb.create_sheet("Data")
        
        # Get columns
        columns = config.get('columns', list(data[0].keys()) if data else [])
        
        # Write headers
        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, "")
                # Convert complex types to string for Excel compatibility
                if isinstance(value, (list, dict)):
                    value = json.dumps(value)
                elif isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Apply conditional formatting
                if col_name == 'severity':
                    self._apply_severity_formatting(cell, value)
                elif col_name == 'status':
                    self._apply_status_formatting(cell, value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook, data: List[Dict[str, Any]], config: Dict[str, Any]):
        """Create summary statistics sheet in Excel."""
        ws = wb.create_sheet("Summary")
        
        # Calculate summary statistics
        total_records = len(data)
        severity_counts = {}
        status_counts = {}
        agency_counts = {}
        
        for row in data:
            severity = row.get('severity', 'Unknown')
            status = row.get('status', 'Unknown')
            agency = row.get('agency_name', 'Unknown')
            
            severity_counts[severity] = severity_counts.get(severity, 0) + 1
            status_counts[status] = status_counts.get(status, 0) + 1
            agency_counts[agency] = agency_counts.get(agency, 0) + 1
        
        # Write summary
        ws['A1'] = "Compliance Monitoring Summary"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Basic stats
        ws['A3'] = "Total Records:"
        ws['B3'] = total_records
        ws['A4'] = "Export Date:"
        ws['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Severity breakdown
        ws['A6'] = "Severity Breakdown"
        ws['A6'].font = Font(bold=True)
        row = 7
        for severity, count in severity_counts.items():
            ws[f'A{row}'] = severity
            ws[f'B{row}'] = count
            row += 1
        
        # Status breakdown
        ws['D6'] = "Status Breakdown"
        ws['D6'].font = Font(bold=True)
        row = 7
        for status, count in status_counts.items():
            ws[f'D{row}'] = status
            ws[f'E{row}'] = count
            row += 1
    
    def _create_charts_sheet(self, wb: openpyxl.Workbook, data: List[Dict[str, Any]], config: Dict[str, Any]):
        """Create charts sheet in Excel."""
        ws = wb.create_sheet("Charts")
        
        # Create severity pie chart
        if data:
            severity_counts = {}
            for row in data:
                severity = row.get('severity', 'Unknown')
                severity_counts[severity] = severity_counts.get(severity, 0) + 1
            
            # Add chart data
            ws['A1'] = "Severity"
            ws['B1'] = "Count"
            row = 2
            for severity, count in severity_counts.items():
                ws[f'A{row}'] = severity
                ws[f'B{row}'] = count
                row += 1
            
            # Create pie chart
            chart = openpyxl.chart.PieChart()
            chart.title = "Changes by Severity"
            data = Reference(ws, min_col=2, min_row=1, max_row=len(severity_counts) + 1)
            labels = Reference(ws, min_col=1, min_row=2, max_row=len(severity_counts) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            ws.add_chart(chart, "D2")
    
    def _create_metadata_sheet(self, wb: openpyxl.Workbook, config: Dict[str, Any]):
        """Create metadata sheet in Excel."""
        ws = wb.create_sheet("Metadata")
        
        ws['A1'] = "Export Metadata"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Export information
        ws['A3'] = "Export Date:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws['A4'] = "Export Format:"
        ws['B4'] = "Excel"
        
        ws['A5'] = "Filters Applied:"
        filters = config.get('filters', {})
        ws['B5'] = json.dumps(filters, indent=2) if filters else "None"
        
        ws['A6'] = "Columns Included:"
        columns = config.get('columns', [])
        ws['B6'] = ", ".join(columns) if columns else "All"
    
    def _create_pdf_metadata(self, styles, config: Dict[str, Any]) -> List:
        """Create PDF metadata section."""
        story = []
        
        # Export information
        metadata_style = ParagraphStyle(
            'Metadata',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=6
        )
        
        export_date = f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        story.append(Paragraph(export_date, metadata_style))
        
        filters = config.get('filters', {})
        if filters:
            filters_text = f"Filters: {json.dumps(filters, indent=2)}"
            story.append(Paragraph(filters_text, metadata_style))
        
        return story
    
    def _create_pdf_summary(self, data: List[Dict[str, Any]], styles) -> List:
        """Create PDF summary section."""
        story = []
        
        # Summary statistics
        total_records = len(data)
        severity_counts = {}
        for row in data:
            severity = row.get('severity', 'Unknown')
            severity_counts[severity] = severity_counts.get(severity, 0) + 1
        
        summary_style = ParagraphStyle(
            'Summary',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12
        )
        
        story.append(Paragraph("Summary Statistics", summary_style))
        
        # Create summary table
        summary_data = [
            ["Metric", "Value"],
            ["Total Records", str(total_records)],
            ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        for severity, count in severity_counts.items():
            summary_data.append([f"{severity} Changes", str(count)])
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        return story
    
    def _create_pdf_data_table(self, data: List[Dict[str, Any]], config: Dict[str, Any], styles) -> List:
        """Create PDF data table section."""
        story = []
        
        if not data:
            return story
        
        # Get columns
        columns = config.get('columns', list(data[0].keys()) if data else [])
        
        # Create table data
        table_data = [columns]  # Header row
        
        for row in data:
            formatted_row = []
            for col in columns:
                value = row.get(col, "")
                # Format datetime values
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                formatted_row.append(str(value))
            table_data.append(formatted_row)
        
        # Limit rows for PDF (first 50 rows)
        table_data = table_data[:51]
        
        # Create table
        data_table = Table(table_data)
        data_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white])
        ]))
        
        story.append(Paragraph("Data Export", styles['Heading2']))
        story.append(data_table)
        
        if len(data) > 50:
            story.append(Paragraph(f"Note: Showing first 50 of {len(data)} records", styles['Normal']))
        
        return story
    
    def _format_row_for_csv(self, row: Dict[str, Any]) -> Dict[str, Any]:
        """Format row data for CSV export."""
        formatted_row = {}
        for key, value in row.items():
            if isinstance(value, datetime):
                formatted_row[key] = value.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(value, dict):
                formatted_row[key] = json.dumps(value)
            else:
                formatted_row[key] = str(value) if value is not None else ""
        return formatted_row
    
    def _apply_severity_formatting(self, cell, value):
        """Apply conditional formatting based on severity."""
        if value == 'critical':
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        elif value == 'high':
            cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        elif value == 'medium':
            cell.fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
        elif value == 'low':
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    
    def _apply_status_formatting(self, cell, value):
        """Apply conditional formatting based on status."""
        if value == 'pending':
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        elif value == 'completed':
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        elif value == 'failed':
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)


class ExportScheduler:
    """Manages scheduled exports and automated delivery."""
    
    def __init__(self, export_manager: ExportManager):
        self.export_manager = export_manager
        self.scheduled_exports = {}
    
    def schedule_export(
        self,
        export_id: str,
        schedule_config: Dict[str, Any],
        export_config: Dict[str, Any]
    ) -> bool:
        """
        Schedule a recurring export.
        
        Args:
            export_id: Unique identifier for the export
            schedule_config: Schedule configuration (frequency, recipients, etc.)
            export_config: Export configuration (format, filters, etc.)
            
        Returns:
            True if scheduled successfully
        """
        try:
            self.scheduled_exports[export_id] = {
                'schedule': schedule_config,
                'export_config': export_config,
                'created_at': datetime.now(),
                'last_run': None,
                'next_run': self._calculate_next_run(schedule_config)
            }
            logger.info(f"Scheduled export {export_id} for {schedule_config.get('frequency', 'unknown')}")
            return True
        except Exception as e:
            logger.error(f"Failed to schedule export {export_id}: {str(e)}")
            return False
    
    def _calculate_next_run(self, schedule_config: Dict[str, Any]) -> datetime:
        """Calculate the next run time based on schedule configuration."""
        frequency = schedule_config.get('frequency', 'daily')
        now = datetime.now()
        
        if frequency == 'daily':
            return now + timedelta(days=1)
        elif frequency == 'weekly':
            return now + timedelta(weeks=1)
        elif frequency == 'monthly':
            # Simple monthly calculation
            return now + timedelta(days=30)
        else:
            return now + timedelta(days=1)  # Default to daily
    
    def get_scheduled_exports(self) -> Dict[str, Any]:
        """Get all scheduled exports."""
        return self.scheduled_exports.copy()
    
    def cancel_export(self, export_id: str) -> bool:
        """Cancel a scheduled export."""
        if export_id in self.scheduled_exports:
            del self.scheduled_exports[export_id]
            logger.info(f"Cancelled export {export_id}")
            return True
        return False


# Global export manager instance
export_manager = ExportManager()
export_scheduler = ExportScheduler(export_manager) 